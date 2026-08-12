#!/usr/bin/env python3
"""
Build the Medhour Pharmacy product catalogue from `MEDHOUR PRICE LIST.xlsx`.

Reads the pharmacy's stock export and produces `scripts/catalogue.json`, the
single source of truth used by `scripts/load-catalogue.mjs` (pushes to Supabase)
and by `supabase/seed.sql`.

Price ("Selling Price") and stock ("On Hand") are copied verbatim from the
spreadsheet — never derived, never rounded. Only the *display name* is tidied:
casing, spacing, expanded dosage-form abbreviations and unambiguous spelling
corrections. The untouched spreadsheet name is kept on every product as
`internal_name` so the pharmacy can always reconcile against their own system.

Usage:  python3 scripts/build_catalogue.py
"""

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "MEDHOUR PRICE LIST.xlsx"
OUT_JSON = ROOT / "scripts" / "catalogue.json"
OUT_REVIEW = ROOT / "scripts" / "CATALOGUE-REVIEW.md"

# --------------------------------------------------------------------------
# Categories
# --------------------------------------------------------------------------

CATEGORIES = [
    ("pain-inflammation", "Pain & Anti-inflammatory", "PAI",
     "Painkillers, anti-inflammatories, muscle rubs and pain-relief gels."),
    ("antibiotics", "Antibiotics & Anti-infectives", "ABX",
     "Antibiotics, antifungals and antiviral medicines."),
    ("malaria", "Malaria Treatment", "MAL",
     "Antimalarial tablets, suspensions and injections."),
    ("cough-cold-allergy", "Cough, Cold & Allergy", "CCA",
     "Cough syrups, cold and flu remedies, antihistamines and asthma inhalers."),
    ("stomach-digestive", "Stomach & Digestive", "DIG",
     "Antacids, ulcer treatment, anti-nausea, laxatives and rehydration."),
    ("diabetes-blood-pressure", "Diabetes & Blood Pressure", "DBP",
     "Diabetes, blood pressure, heart and cholesterol medicines and test strips."),
    ("skin-care", "Skin Care & Creams", "SKN",
     "Creams, ointments and lotions for skin conditions."),
    ("eye-ear-oral", "Eye, Ear & Oral Care", "EEO",
     "Eye drops, ear drops, mouthwash and dental care."),
    ("vitamins-supplements", "Vitamins & Supplements", "VIT",
     "Vitamins, minerals, tonics and food supplements."),
    ("deworming", "Deworming", "DWM",
     "Albendazole, mebendazole and other deworming medicines."),
    ("sexual-reproductive-health", "Sexual & Reproductive Health", "SRH",
     "Condoms, contraceptives, fertility and intimate health products."),
    ("mother-baby", "Mother & Baby", "MOB",
     "Prenatal care, baby syrups, baby skincare and delivery kits."),
    ("injections-iv", "Injections & IV Fluids", "INJ",
     "Injectable medicines, IV fluids and vaccines."),
    ("medical-supplies", "Medical Supplies & Devices", "SUP",
     "Syringes, gloves, cotton wool, cannulas, test strips and blades."),
    ("first-aid", "First Aid & Antiseptics", "FAD",
     "Antiseptics, wound care, burn creams and first-aid essentials."),
    ("personal-care", "Personal Care", "PER",
     "Everyday personal, skin and household health care."),
    ("general-medicines", "General Medicines", "GEN",
     "Other prescription and general medicines held in stock."),
]

CATEGORY_CODES = {slug: code for slug, _n, code, _d in CATEGORIES}

# --------------------------------------------------------------------------
# Spelling corrections — applied as whole words, case-insensitively.
# Only unambiguous misspellings of *generic drug names* and English words.
# Brand names are left alone (they are simply title-cased).
# --------------------------------------------------------------------------

SPELLING = {
    # generic drug names
    "pracetamol": "Paracetamol", "paracetmol": "Paracetamol",
    "ctriaxone": "Ceftriaxone", "ondesatron": "Ondansetron",
    "ondensetron": "Ondansetron", "floconazole": "Fluconazole",
    "capropril": "Captopril", "caropril": "Captopril",
    "chromoglycate": "Cromoglicate",
    "contrimoxazole": "Co-trimoxazole", "cotrimoxazole": "Co-trimoxazole",
    "diclifenac": "Diclofenac", "dicloifenac": "Diclofenac",
    "metrgonidazole": "Metronidazole", "metromidazole": "Metronidazole",
    "metrondazole": "Metronidazole",
    "carbamezapin": "Carbamazepine", "carbamezapine": "Carbamazepine",
    "phenobabiton": "Phenobarbitone", "nitrofurantion": "Nitrofurantoin",
    "nofloxacim": "Norfloxacin", "asprin": "Aspirin", "piroxcam": "Piroxicam",
    "prednisolon": "Prednisolone", "latulose": "Lactulose",
    "acyclivir": "Acyclovir", "gentamycion": "Gentamicin",
    "gentamycin": "Gentamicin", "soldenafil": "Sildenafil",
    "mometasine": "Mometasone", "desloratasine": "Desloratadine",
    "hydrocortsone": "Hydrocortisone", "betamethason": "Betamethasone",
    "clotrimazol": "Clotrimazole", "xclotrimazole": "Clotrimazole",
    "bromheine": "Bromhexine", "artesurnate": "Artesunate",
    "akbendazole": "Albendazole", "multvitamin": "Multivitamin",
    "aminoflyne": "Aminophylline", "domperidon": "Domperidone",
    "levocetirizen": "Levocetirizine", "cetrizen": "Cetirizine",
    "cetrizine": "Cetirizine", "cetirizen": "Cetirizine",
    "atovastatin": "Atorvastatin", "amlodipin": "Amlodipine",
    "candesatan": "Candesartan", "candesatrtan": "Candesartan",
    "doxycline": "Doxycycline", "amoxycilin": "Amoxycillin",
    "chlorampenicol": "Chloramphenicol", "clarythromycin": "Clarithromycin",
    "intraconazole": "Itraconazole", "mtero": "Metronidazole",
    "lidinocain": "Lidocaine", "micona": "Miconazole",
    "ofloxacin": "Ofloxacin", "albendazole": "Albendazole",
    "mebendazole": "Mebendazole", "clomiphene": "Clomiphene",
    # english words
    "cottol": "Cotton", "gaurd": "Guard", "gard": "Guard", "watar": "Water",
    "perixide": "Peroxide", "potasium": "Potassium",
    "pamangament": "Permanganate", "vacejne": "Vaccine",
    "troberry": "Strawberry", "strowberry": "Strawberry",
    "pinaple": "Pineapple", "listerin": "Listerine",
    "crem": "Cream", "cdream": "Cream", "kream": "Cream",
    "conndom": "Condom", "condom": "Condom", "clarytine": "Claritine",
    "bisacodly": "Bisacodyl", "chlorampenicol": "Chloramphenicol",
    "mucogelbsyrup": "Mucogel Syrup", "ttoff": "Toff",
    "claryhthromycin": "Clarithromycin", "cirpokant": "Ciprokant",
    "haemorrhodial": "Haemorrhoidal", "clobetasole": "Clobetasol",
    "sandox": "Sandoz", "shels": "Shelys", "ellys": "Elys", "elly": "Elys",
    "cypus": "Cyprus", "ajant": "Ajanta", "grwmin": "Growmin",
    "peppeerrmint": "Peppermint", "metochlopramide": "Metoclopramide",
    "metoclopamid": "Metoclopramide", "belladona": "Belladonna",
    "tramadiol": "Tramadol", "diclopara": "Diclopara",
    "whitefield": "Whitfield's", "asprine": "Aspirin",
    "syrp": "Syrup", "syru": "Syrup", "sryp": "Syrup", "ayrup": "Syrup",
    "syp": "Syrup", "suap": "Syrup", "srup": "Syrup", "syrup": "Syrup",
    "cought": "Cough", "coughtt": "Cough",
    "peppeermint": "Peppermint", "strowbery": "Strawberry",
    "viginal": "Vaginal", "vagina": "Vaginal",
    "supp": "Suppositories", "pess": "Pessaries",
    "freez": "Freeze", "vomistant": "Vomistant",
}

# Dosage-form abbreviations -> expanded word.
FORMS = {
    "tbs": "Tablets", "tab": "Tablets", "tabs": "Tablets", "tablet": "Tablets",
    "tablets": "Tablets", "tbls": "Tablets",
    "caps": "Capsules", "cap": "Capsules", "capsule": "Capsules",
    "capsules": "Capsules", "softygel": "Softgel",
    "susp": "Suspension", "suspension": "Suspension", "sus": "Suspension",
    "inj": "Injection", "injection": "Injection",
    "oint": "Ointment", "ont": "Ointment", "ointment": "Ointment",
    "sol": "Solution", "solution": "Solution",
    "disp": "Disposable",
    "pead": "Paediatric", "ped": "Paediatric", "paed": "Paediatric",
    "lozenge": "Lozenges", "lozenges": "Lozenges",
    "amp": "Ampoule",
}

# Tokens that must stay upper-case when title-casing.
KEEP_UPPER = {
    "IV", "ORS", "UPT", "DNS", "IU", "BP", "HC", "DT", "OD", "VP", "DG",
    "MR", "SF", "DM", "P2", "GSK", "V6", "V3", "V2", "V", "B", "G", "GM",
    "C", "D", "E", "K", "N", "P", "S", "MX", "AA", "KLY", "RB", "BBE",
    "II", "SKF", "TZ", "USI", "LCT", "CT", "AZ", "PA12", "IPCA", "DR",
}

# Unit tokens — lower-cased only when they directly follow a number.
UNIT_TOKENS = {"ml", "mg", "g", "gm", "cc", "cm", "kg", "mcg"}

# Connectives that must stay lower-case.
KEEP_LOWER = {"x", "ya", "na", "de", "of", "and", "with", "for"}


def presplit(text: str) -> str:
    """Separate run-together tokens before word-level replacement.

    "375mgTbs" -> "375mg Tbs", "100sCYPRUS" -> "100s CYPRUS", "10;X10" -> "10 X10"
    """
    t = text.replace(";", " ")
    # Split only when a real word follows the unit, so "500gm" is not read as
    # "500g" + "m" and "100mls" is not read as "100ml" + "s".
    t = re.sub(r"(\d\s?(?:mgs|mg|mls|ml|gms|gm|g|iu|cc|cm))([A-Z][a-zA-Z]{2,})",
               r"\1 \2", t)
    t = re.sub(r"(\d+s)([A-Z])", r"\1 \2", t)
    t = re.sub(r"(\d)\s*[xX][sS](\d)", r"\1 x \2", t)
    return re.sub(r"\s+", " ", t)


def fix_units(text: str) -> str:
    """Normalise number+unit spacing without changing any numeric value."""
    t = text
    t = re.sub(r"(\d)\s*mg\s*mg\b", r"\1 mg", t, flags=re.I)
    t = re.sub(r"(\d)\s*(?:mls|mlos|mks|ls)\b", r"\1 ml", t, flags=re.I)
    t = re.sub(r"(\d)\s*ml\b", r"\1 ml", t, flags=re.I)
    t = re.sub(r"(\d)\s*(?:mgs|mg)\b", r"\1 mg", t, flags=re.I)
    t = re.sub(r"(\d)\s*(?:gms|gm|g)\b", r"\1 g", t, flags=re.I)
    t = re.sub(r"(\d)\s*IU\b", r"\1 IU", t, flags=re.I)
    t = re.sub(r"(\d)\s*cc\b", r"\1 cc", t, flags=re.I)
    t = re.sub(r"(\d)\s*cm\b", r"\1 cm", t, flags=re.I)
    t = re.sub(r"(\d)\s*[xX]\s*(\d)", r"\1 x \2", t)
    t = re.sub(r"(\d)\s*/\s*(\d)", r"\1/\2", t)
    t = re.sub(r"(\d)\s*'?[sS]\b", r"\1s", t)
    # collapse a doubled unit produced by source typos ("500MGmg" -> "500 mg")
    t = re.sub(r"\b(mg|ml|g)\s+\1\b", r"\1", t, flags=re.I)
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def titlecase_token(tok: str, prev: str = "") -> str:
    bare = tok.strip("()[].,")
    if not bare:
        return tok
    low = bare.lower()
    # a unit only counts as a unit when it follows a number
    if low in UNIT_TOKENS:
        if prev and prev.rstrip("()[].,%").endswith(tuple("0123456789")):
            return tok.replace(bare, low)
    if low in KEEP_LOWER:
        return tok.replace(bare, low)
    if bare.upper() in KEEP_UPPER and bare.isalpha():
        return tok.replace(bare, bare.upper())
    if re.match(r"^\d", bare):
        return tok
    if re.match(r"^[A-Za-z]+$", bare):
        return tok.replace(bare, bare[0].upper() + bare[1:].lower())
    # "EYE/EAR" -> "Eye/Ear", "D-ARTEPP" -> "D-Artepp"; leave mixed-case
    # tokens ("IU/ml", "V6") alone.
    for sep in ("/", "-"):
        if bare.isupper() and re.match(rf"^[A-Za-z]+{re.escape(sep)}[A-Za-z]+$", bare):
            parts = [p if p in KEEP_UPPER else p.capitalize() for p in bare.split(sep)]
            return tok.replace(bare, sep.join(parts))
    return tok.replace(bare, bare.upper() if bare.isupper() else bare)


def clean_name(raw: str) -> str:
    t = presplit(raw.strip())
    # balance brackets left unclosed / unopened in the spreadsheet
    if t.count(")") > t.count("("):
        t = t.replace(")", "", t.count(")") - t.count("("))
    elif t.count("(") > t.count(")"):
        t = t + ")" * (t.count("(") - t.count(")"))
    t = re.sub(r"\s+", " ", t).strip(" -,")

    # word-level corrections
    out = []
    for tok in t.split(" "):
        lead = ""
        trail = ""
        m = re.match(r"^([(\[]*)(.*?)([)\].,;]*)$", tok)
        if m:
            lead, core, trail = m.group(1), m.group(2), m.group(3)
        else:
            core = tok
        low = core.lower()
        if low in SPELLING:
            core = SPELLING[low]
        elif low in FORMS:
            core = FORMS[low]
        out.append(lead + core + trail)
    t = " ".join(out)

    # "SUP"/"SUPP" means suspension when followed by a millilitre volume
    t = re.sub(r"\bSUPP?\b(?=\s*\d+\s*ml)", "Suspension", t, flags=re.I)
    t = re.sub(r"\bSUPP?\s*(\d+\s*ml)", r"Suspension \1", t, flags=re.I)
    t = re.sub(r"\bSUPP?\b", "Suppositories", t, flags=re.I)

    t = fix_units(t)
    tokens = t.split(" ")
    cased = []
    for i, tok in enumerate(tokens):
        cased.append(titlecase_token(tok, tokens[i - 1] if i else ""))
    t = " ".join(cased)
    t = re.sub(r"\s+", " ", t).strip(" -,")
    t = t.replace("( ", "(").replace(" )", ")")
    return t


# --------------------------------------------------------------------------
# Dosage form detection
# --------------------------------------------------------------------------

FORM_PATTERNS = [
    ("Injection", r"\b(inj|injection|iv|ampoule|amp)\b|\bvaccine\b"),
    ("Inhaler", r"\binhaler\b"),
    ("Eye/Ear drops", r"\b(eye|ear)\b.*\b(drop|ointment|oint)\b|\bdrop\b.*\b(eye|ear)\b"),
    ("Drops", r"\bdrops?\b"),
    ("Syrup", r"\bsyrup\b"),
    ("Suspension", r"\b(susp|suspensions?|sus)\b"),
    ("Pessaries", r"\bpess\w*"),
    ("Suppositories", r"\bsupp\w*"),
    ("Cream", r"\bcream\b"),
    ("Ointment", r"\b(oint|ointment|ont)\b"),
    ("Gel", r"\bgel\b|\bjel\b"),
    ("Lotion", r"\blotion\b"),
    ("Shampoo", r"\bshampoo\b"),
    ("Mouthwash", r"mouth\s*wash"),
    ("Lozenges", r"lozenge"),
    ("Powder", r"\bpowder\b"),
    ("Sachet", r"\bsachets?\b"),
    ("Solution", r"\b(sol|solution|liquid|mixture|liniment|tincture|spirit)\b"),
    ("Spray", r"\bspray\b"),
    ("Oil", r"\boil\b"),
    ("Balm", r"\bbalm\b"),
    ("Capsules", r"\b(caps?|capsules?|softgel)\b"),
    ("Tablets", r"\b(tbs|tabs?|tablets?|kit)\b"),
]


def detect_form(raw: str) -> str | None:
    low = raw.lower()
    for form, pattern in FORM_PATTERNS:
        if re.search(pattern, low):
            return form
    return None


# How a countable pack is worded per dosage form.
COUNT_WORD = {
    "Tablets": "tablet", "Capsules": "capsule", "Pessaries": "pessary",
    "Suppositories": "suppository", "Lozenges": "lozenge",
    "Sachet": "sachet", "Injection": "ampoule",
}

# Forms measured by volume, and forms measured by tube weight.
VOLUME_FORMS = {
    "Syrup", "Suspension", "Solution", "Drops", "Eye/Ear drops", "Mouthwash",
    "Lotion", "Shampoo", "Oil", "Spray", "Injection",
}
WEIGHT_FORMS = {"Cream", "Ointment", "Gel", "Powder", "Balm"}


PLURALS = {"pessary": "pessaries", "suppository": "suppositories"}


def _plural(word: str) -> str:
    return PLURALS.get(word, word + "s")


def _count(n: str, form: str | None) -> str:
    word = COUNT_WORD.get(form or "", "piece")
    return f"{n} {word}" if n == "1" else f"{n} {_plural(word)}"


def detect_pack(name: str, form: str | None) -> str | None:
    """Derive a human pack size from the cleaned display name.

    Reads only *counts* and *container sizes* — never a drug strength. A
    strength always carries a unit (mg / g / ml) attached to the active
    ingredient, so form drives which pattern is trusted.
    """
    grid = re.search(r"(\d+)\s*x\s*(\d+)", name, flags=re.I)
    count = re.search(r"(\d+)\s*s\b", name)
    volume = re.search(r"(\d+)\s*ml\b", name, flags=re.I)
    weight = re.search(r"(\d+)\s*g\b", name, flags=re.I)
    cc = re.search(r"(\d+)\s*cc\b", name, flags=re.I)

    if form in VOLUME_FORMS:
        if volume:
            return f"{volume.group(1)} ml"
        if form == "Injection" and count:
            return _count(count.group(1), form)
        return None

    if form in WEIGHT_FORMS:
        if weight:
            return f"{weight.group(1)} g"
        if volume:
            return f"{volume.group(1)} ml"
        return None

    # tablets, capsules, pessaries, kits, devices, unknown
    if grid:
        word = COUNT_WORD.get(form or "", "piece")
        return f"{grid.group(1)} x {grid.group(2)} {_plural(word)}"
    if count:
        return _count(count.group(1), form)
    if cc:
        return f"{cc.group(1)} cc"
    if volume:
        return f"{volume.group(1)} ml"
    if weight and form not in {"Tablets", "Capsules"}:
        return f"{weight.group(1)} g"
    return None


# --------------------------------------------------------------------------
# Generic-name detection (only confident, unambiguous matches)
# --------------------------------------------------------------------------

GENERIC_RULES = [
    (r"\bparacetamol|\bpanadol|\bcalpol\b|\belymol\b|\bzenadol\b|\bdolomol\b|\bcetamol\b|\bsheladol\b|\btotomol\b|\bpa\s?12\b|\badol\b|\bprinadol\b", "Paracetamol"),
    (r"\bibuprofen|\bibubeg\b|\bibumex\b|\bgofen\b|\bibun\b", "Ibuprofen"),
    (r"\bdiclofenac|\bdiclokant\b|\bdicloran\b|\bdicloday\b|\bdiclopar\b|\brelaxo gel\b|\brehumac\b|\bremethan\b|\bzerodol\b|\bzeradol\b|\bvivian\b|\bvolin gel\b", "Diclofenac"),
    (r"\baceclofenac|\bacefen\b", "Aceclofenac"),
    (r"\bmefenamic", "Mefenamic acid"),
    (r"\bpiroxicam", "Piroxicam"),
    (r"\bmuvera\b|\bmeloxicam", "Meloxicam"),
    (r"\btramadol|\bdomadol\b|\btramazac\b", "Tramadol"),
    (r"\bketoprofen|\bfastum\b|\bketogesic\b", "Ketoprofen"),
    (r"\baspirin|\bascard\b", "Aspirin"),
    (r"\bamoxycillin|\bamoxicillin|\bamoxycilin", "Amoxycillin"),
    (r"\bampiclox|\bampicillin", "Ampicillin + Cloxacillin"),
    (r"\bco-?amoxiclav|\bclavam\b|\bclav\b|\bclavulan|clav \d|\bcledomox\b|\bkoact\b|\bredmentin\b|\bindiclav\b|\balphaclav\b|\bmyclav\b|\bsportclav\b|\bspotclav\b", "Amoxycillin + Clavulanic acid"),
    (r"\bciprofloxacin|\bcipro\b|\bciprokant\b|\bzindolin\b|\bcirpokant\b", "Ciprofloxacin"),
    (r"\bofloxacin|\bpinox\b|\btoflox\b", "Ofloxacin"),
    (r"\bnorfloxacin", "Norfloxacin"),
    (r"\blevofloxacin|\blevoz\b", "Levofloxacin"),
    (r"\bazithromycin|\bazilin\b|\bazecure\b|\bazithraa\b|\bazuma\b|\bzaha\b|\bagycin\b", "Azithromycin"),
    (r"\berythromycin", "Erythromycin"),
    (r"\bclarithromycin|\bclarie\b|\bclaranta\b", "Clarithromycin"),
    (r"\bclindamycin|\btidact\b", "Clindamycin"),
    (r"\bdoxycycline|\bdoxy\b", "Doxycycline"),
    (r"\btetracycline", "Tetracycline"),
    (r"\bchloramphenicol|\bcomycetin\b|\bprinaton\b", "Chloramphenicol"),
    (r"\bco-trimoxazole|\bprinatrim\b|\bseptrin\b", "Co-trimoxazole"),
    (r"\bmetronidazole|\bbetrozole\b|\bprinalyn\b|\binflazole\b", "Metronidazole"),
    (r"\btinidazole", "Tinidazole"),
    (r"\bsecnidazole|\bdysen\b", "Secnidazole"),
    (r"\bornidazole|\borgly\b", "Ornidazole"),
    (r"\bcefixime|\bcefibac\b|\bsanix\b|\bfixinet\b|\binofix\b|\bc-tax\b", "Cefixime"),
    (r"\bcefuroxime|\bauxtocef\b", "Cefuroxime"),
    (r"\bcefadroxil|\bdrox\b", "Cefadroxil"),
    (r"\bcephalexin", "Cephalexin"),
    (r"\bceftriaxone|\baksone\b", "Ceftriaxone"),
    (r"\bflucloxacillin|\bflucamox\b", "Flucloxacillin"),
    (r"\bpen v\b|\bphenoxymethylpenicillin", "Phenoxymethylpenicillin"),
    (r"\bnitrofurantoin", "Nitrofurantoin"),
    (r"\bgentamicin|\bgentabac\b", "Gentamicin"),
    (r"\bnystatin", "Nystatin"),
    (r"\bfluconazole|\bzocon\b|\bfluderm\b|\bflucoz\b", "Fluconazole"),
    (r"\bitraconazole|\bcanditral\b", "Itraconazole"),
    (r"\bgriseofulvin", "Griseofulvin"),
    (r"\bterbinafine|\bbinafin\b", "Terbinafine"),
    (r"\bketoconazole|\bketoneal\b|\bketineal\b|\bketoz\b|\bdermazole\b", "Ketoconazole"),
    (r"\bclotrimazole|\bcandid\b|\bcanestal\b|\bcandistant\b|\blabesten\b|\bfuncream\b|\bfunkream\b|\bkandecide\b|\bcanimask\b|\bprocostant\b|\bpricostant\b|\btranspol\b|\btransipol\b", "Clotrimazole"),
    (r"\bmiconazole|\bfungarin\b|\bgynozole\b|\bmicor\b", "Miconazole"),
    (r"\bacyclovir|\baciclovir", "Acyclovir"),
    (r"\bmupirocin|\bbactroban|\bbactopic|\bbactrokant\b|\bpustalrocin\b|\bplustalrocin\b|\bemina\b", "Mupirocin"),
    (r"\bartemether|\bartefan\b|\bartefen\b|\bartefeb\b|\blumerax\b|\blartem\b|\blaterm\b|\blornat\b|\bmalafin\b|\bd-artepp\b", "Artemether + Lumefantrine"),
    (r"\bartesunate|\bartesun\b", "Artesunate"),
    (r"\bartequick\b|\bduocotexin\b|\bdihydroartemisinin", "Dihydroartemisinin + Piperaquine"),
    (r"\balbendazole|\bzental\b|\bzentel\b|\banthel\b|\bwomiban\b|\bfilazole\b|\balben\b|\belyzole\b|\bellyzole\b|\balbasol\b", "Albendazole"),
    (r"\bmebendazole|\bworminil\b|\bnatoa\b", "Mebendazole"),
    (r"\bcetirizine|\boncet\b|\bzirin\b|\bsyneez\b|\bsancet\b|\bejkon\b|\bsatrin\b|\ballerid\b", "Cetirizine"),
    (r"\blevocetirizine|\blevozin\b", "Levocetirizine"),
    (r"\bloratadine|\bloratyn\b|\blorhistina\b|\blorata\b|\btidilor\b|\blara\b|\bclaritine\b|\bclarityne\b|\batadyn\b", "Loratadine"),
    (r"\bdesloratadine|\bglendes\b", "Desloratadine"),
    (r"\bchlorphenamine|\bchlorpheniramine", "Chlorphenamine"),
    (r"\bpromethazine", "Promethazine"),
    (r"\bsalbutamol|\basthalin\b|\basthamol\b", "Salbutamol"),
    (r"\bmontelukast|\bkipel\b|\bmontemac\b", "Montelukast"),
    (r"\baminophylline", "Aminophylline"),
    (r"\bbudesonide|\bbudecort\b", "Budesonide"),
    (r"\bambroxol|\bambrodil\b|\bambrox\b|\babrox\b|\bambrosan\b", "Ambroxol"),
    (r"\bbromhexine|\bmucospel\b", "Bromhexine"),
    (r"\bomeprazole|\bomesk\b|\bzosec\b", "Omeprazole"),
    (r"\besomeprazole|\besoz\b", "Esomeprazole"),
    (r"\bpantoprazole|\bpanacid\b|\bpantonex\b|\bpatmac\b", "Pantoprazole"),
    (r"\brabeprazole|\brabeloc\b", "Rabeprazole"),
    (r"\blansoprazole|\bzolanas\b", "Lansoprazole"),
    (r"\bsucralfate|\bsucrafil\b", "Sucralfate"),
    (r"\bdomperidone|\bmotinorm\b|\bmomi\b", "Domperidone"),
    (r"\bmetoclopramide|\bperinorm\b", "Metoclopramide"),
    (r"\bondansetron|\bvomikind\b|\bvomistant\b", "Ondansetron"),
    (r"\bhyoscine|\bbisponol\b", "Hyoscine butylbromide"),
    (r"\bloperamide|\blopa\b", "Loperamide"),
    (r"\blactulose|\blaxalink\b|\blacteas\b", "Lactulose"),
    (r"\bbisacodyl", "Bisacodyl"),
    (r"\bsimethicone|\bcoliza\b", "Simethicone"),
    (r"\bmetformin|\bilet\b", "Metformin"),
    (r"\bglibenclamide|\bgliben\b|\bglitisol\b|\bdiamide\b|\bglucored\b", "Glibenclamide"),
    (r"\bglimepiride|\bglimiprede\b|\bgemar\b", "Glimepiride"),
    (r"\bpioglitazone|\bpio safe\b", "Pioglitazone"),
    (r"\binsulin|\binsulatard\b", "Insulin"),
    (r"\bamlodipine|\bswamlo\b|\bcalchek\b|\basomex\b", "Amlodipine"),
    (r"\bbisoprolol|\bbisotrol\b", "Bisoprolol"),
    (r"\batenolol|\batenelo\b", "Atenolol"),
    (r"\bcaptopril", "Captopril"),
    (r"\bmethyldopa", "Methyldopa"),
    (r"\bnifedipine", "Nifedipine"),
    (r"\blosartan|\bpresartan\b|\bpresatan\b", "Losartan"),
    (r"\bcandesartan|\bcandez\b|\baderan\b", "Candesartan"),
    (r"\batorvastatin|\batorem\b", "Atorvastatin"),
    (r"\bclopidogrel|\bclopact\b", "Clopidogrel"),
    (r"\bfrusemide|\bfurosemide", "Furosemide"),
    (r"\bsildenafil|\bkamagra\b|\bsilden\b|\bsilmelt\b|\berector\b", "Sildenafil"),
    (r"\btadalafil|\bmegalis\b|\bapcalis\b", "Tadalafil"),
    (r"\btamsulosin|\burimax\b", "Tamsulosin"),
    (r"\bdiazepam", "Diazepam"),
    (r"\blorazepam|\blorivan\b", "Lorazepam"),
    (r"\bcarbamazepine", "Carbamazepine"),
    (r"\bphenobarbitone|\bphenobarbital", "Phenobarbitone"),
    (r"\bchlorpromazine", "Chlorpromazine"),
    (r"\bhaloperidol", "Haloperidol"),
    (r"\bbaclofen", "Baclofen"),
    (r"\bpregabalin|\bpregasafe\b", "Pregabalin"),
    (r"\bprednisolone", "Prednisolone"),
    (r"\bdexamethasone|\bx-sone\b|\bdexiwn\b", "Dexamethasone"),
    (r"\bbetamethasone|\belyvate\b|\btasone\b|\bbetaderm\b|\bbetad\b|\bdiprosalic\b|\bdiprofos\b", "Betamethasone"),
    (r"\bhydrocortisone|\belycort\b|\bhycorum\b", "Hydrocortisone"),
    (r"\bmometasone|\bmomate\b|\bmomaderm\b|\belocom\b", "Mometasone"),
    (r"\bclobetasol|\bpowercort\b|\bclobederm\b", "Clobetasol"),
    (r"\btranexamic|\btrantum\b|\bfibronex\b|\btransic\b", "Tranexamic acid"),
    (r"\bmisoprostol|\bmisoprost\b", "Misoprostol"),
    (r"\bdydrogesterone|\bduphaston\b", "Dydrogesterone"),
    (r"\bclomiphene|\bclomitab\b", "Clomiphene citrate"),
    (r"\blevonorgestrel|\bp2\b|\bemerginor\b|\bpostinor\b|\bnjoi\b", "Levonorgestrel"),
    (r"\bhydroxyurea", "Hydroxyurea"),
    (r"\bfolic acid|\bferolic\b", "Folic acid"),
    (r"\bnitrofurantoin", "Nitrofurantoin"),
    (r"\blignocaine|\blidocaine", "Lignocaine"),
    (r"\btimolol|\blotim\b", "Timolol"),
    (r"\blatanoprost|\bglauconil\b", "Latanoprost"),
    (r"\bcromoglicate", "Sodium cromoglicate"),
    (r"\bpovidone iodine|\bpyodine\b", "Povidone iodine"),
    (r"\bhydrogen peroxide", "Hydrogen peroxide"),
    (r"\bcalamine", "Calamine"),
    (r"\bbenzoyl peroxide|\bbenzox\b|\bbenox\b|\bperso\b", "Benzoyl peroxide"),
    (r"\bsilver sulphadiazine|\bsilverex\b|\bsilver kant\b|\bargiseptic\b", "Silver sulphadiazine"),
    (r"\boral rehydration|\bors\b", "Oral rehydration salts"),
    (r"\bvitamin c|\bascorbic|\bbiocee\b", "Vitamin C"),
    (r"\bvitamin b complex|\bbecoshel\b|\bneuro-? ?forte\b", "Vitamin B complex"),
    (r"\bvitamin e\b|\benat\b", "Vitamin E"),
    (r"\bmultivitamin|\bvitacap\b", "Multivitamin"),
    (r"\bzinc\b", "Zinc"),
    (r"\bcalcium|\bcalcivita\b|\bcaalvita\b", "Calcium"),
    (r"\bnimesulide|\bnosic\b", "Nimesulide"),
    (r"\bneomycin", "Neomycin"),
    (r"\bwater for injection", "Water for injection"),
]


def detect_generic(raw: str) -> str | None:
    low = raw.lower()
    for pattern, generic in GENERIC_RULES:
        if re.search(pattern, low):
            return generic
    return None


# --------------------------------------------------------------------------
# Category rules — first match wins, so order matters.
# --------------------------------------------------------------------------

CATEGORY_RULES = [
    # --- explicitly non-medicine supplies first -------------------------
    ("medical-supplies", r"\bsyringe|\bcannula\b|\bgloves\b|\bcotton wool\b|surgical blade|\bmackintosh\b|\bupt\b|glucoplus|glucose.*strip|\bstrip\b.*\d+s|\bthermometer\b|\bnebul"),
    ("sexual-reproductive-health", r"condom|\bp2\b|\bemerginor\b|\bnjoi\b|trust (daisy|lily)|\bkly gel\b|clomiphene|clomitab|duphaston|misoprost|\berector\b|kamagra|megalis|apcalis|sildenafil|silmelt|silden\b|tadalafil|\bgyno|gynozole|gynex|\bv6\b|\bv3\b|\bv2\b|vaginal|\bpess\b|pessar|labesten|invel forte|delivery mama|\bvagid\b|\bvp\b|\bv\s+(cream|crem|gel|pess|supp|suppositories|pessaries)"),
    ("mother-baby", r"\bbaby\b|gripe water|zincast|\binfant\b|\bprenatal\b|delivery mama|\bkidicare\b"),
    ("deworming", r"albendazole|mebendazole|\bzental\b|\bzentel\b|\banthel\b|\bwomiban\b|filazole|\balben\b|elyzole|ellyzole|albasol|worminil|\bnatoa\b|\bbenpham\b"),
    ("malaria", r"artefan|artefen|artefeb|artesun|artesurnate|lumerax|\blartem\b|\blaterm\b|\blornat\b|d-artepp|artequick|duocotexin|malafin|artemether"),
    # --- injections and IV fluids ---------------------------------------
    ("injections-iv", r"\binj\b|injection|\biv\b|\bamp\b|\bvaccine\b|\bvacejne\b|water for injection|\bdns\b|\bdiprofos\b|\btransic\b|\bpatmac\b"),
    # --- eye / ear / oral ------------------------------------------------
    ("eye-ear-oral", r"eye|\bear\b|mouth ?wash|\bsensodyne\b|\btoothpaste\b|clenora|medi oral|oral care|\bdental\b|glauconil|timolol|cromoglicate|\boralcure\b|\bmicona oral gel\b|\bmicor\b"),
    # --- first aid (before skin: burn creams and antiseptics belong here) --
    ("first-aid", r"antiseptic|povidone|iodine|\bspirit\b|\beusol\b|\bbbe\b|hydrogen peroxide|potassium permanganate|boric acid|\bburn\b|\bburnox\b|silverex|silver kant|argiseptic|calamine|\bbandage\b|\bplaster\b|\bgauze\b|\bwhitefield\b"),
    # --- skin -------------------------------------------------------------
    ("skin-care", r"\bcream\b|\boint\b|ointment\b|\bont\b|\blotion\b|shampoo|\bscaboma\b|acne|\bskederm\b|\bskderm\b|dermovit|dermazole|dermofix|dermaquat|dermidex|dermosporin|\bngoziderm\b|\bsafi\b|\bmedivine\b|\blucin\b|\bkinheal\b|\bentezma\b|\bquadragel\b|\bperso\b|benzox|\bbenox\b|\bmicona\b.*gel|metronidazole (dg )?gel|\bcandid\b|\bbezic\b|\btidact gel\b|\buniosten\b|\bintamin\b|\bdawa ya mba\b|\badacin\b"),
    # --- pain -------------------------------------------------------------
    ("pain-inflammation", r"paracetamol|panadol|\bcalpol\b|\belymol\b|\bzenadol\b|\bdolomol\b|\bcetamol\b|\bsheladol\b|\btotomol\b|\bpa ?12\b|\bprinadol\b|ibuprofen|ibubeg|ibumex|\bgofen\b|\bibun\b|diclofenac|diclokant|dicloran|dicloday|diclopar|\brelaxo\b|rehumac|remethan|zerodol|zeradol|\bvivian\b|\bvolin\b|aceclofenac|\bacefen\b|mefenamic|piroxicam|\bmuvera\b|tramadol|domadol|tramazac|ketoprofen|\bfastum\b|ketogesic|\bflamar\b|\balgic\b|\bdentamol\b|muscle plus|muscoleve|\bnauma balm\b|deep (heat|freeze|freez)|\bbaclofen\b|\bnosic\b|\bsalimia\b"),
    # --- cough / cold / allergy -------------------------------------------
    ("cough-cold-allergy", r"\bcof\b|\bcough\b|\bkof\b|\bcoff\b|cetirizine|loratadine|desloratadine|chlorphenamine|promethazine|salbutamol|asthalin|asthamol|inhaler|montelukast|aminophylline|budecort|aerocort|ambroxol|ambrodil|ambrox|abrox|ambrosan|bromhexine|mucospel|\bzecuf\b|\bstrepsils\b|\bvisking\b|\bcoldril\b|\bdr\.? cold\b|\bsolvin\b|\bascoril\b|\bbenylin\b|\bsediton\b|\bsedikof\b|\btossil\b|\bkofol\b|\bmucolyn\b|\bmucogel\b|\bemdelyn\b|\btotolin\b|\bbrozen\b|\bzentus\b|\bdelease|\bpharmacof|\bpharmactin\b|\bshestcof\b|\bchest cof\b|\bkilikof\b|\bdawakof\b|\bgoodmorning\b|\bjunior care\b|\bsktone\b|\balatro\b|\brohisol\b|\bsatrin\b|\ballerid\b|\bappdly\b|\bsyxten\b|\bephedrine\b|nasal|\babnal\b|menth plus|\bcoldrid\b|\btoff ?plus\b|\btoffplus\b|\bcoffdex\b|\bkofyln\b|\bkoflyn\b|\bclarytine\b|\bflucor day\b|\bvasograin\b|\bkoflem\b|\bcofta\b|\bcoffnil\b|\bsyneez\b|\bsancet\b|\boncet\b|\bzirin\b|\bejkon\b|\blevozin\b|\btidilor\b|\blorata\b|\bclaritine\b|\bclarityne\b|\batadyn\b|\blara tbs\b|\bglendes\b|\bloratyn\b|\blorhistina\b"),
    # --- stomach / digestive ----------------------------------------------
    ("stomach-digestive", r"omeprazole|omesk|\bzosec\b|esomeprazole|\besoz\b|pantoprazole|panacid|pantonex|rabeprazole|rabeloc|lansoprazole|zolanas|sucralfate|sucrafil|antacid|\bactal\b|\beno\b|\balugel\b|\bcorecid\b|altapham|\bnilacid\b|\brelcer\b|gaviscon|\bmagnomint\b|domperidone|motinorm|metoclopramide|perinorm|ondansetron|vomikind|vomistant|hyoscine|bisponol|loperamide|\blopa\b|lactulose|laxalink|lacteas|bisacodyl|osmolax|\bors\b|rehydration|glucose powder|heligo|simethicone|coliza|anusol|haemorrhoid|belladona|belladonna|\bcital\b|\bnetazox\b|\bdysen\b|secnidazole|\bmucogel\b"),
    # --- diabetes / BP / heart --------------------------------------------
    ("diabetes-blood-pressure", r"metformin|\bilet\b|glibenclamide|gliben|glitisol|diamide|glucored|glimepiride|glimiprede|\bgemar\b|pioglitazone|pio safe|insulin|insulatard|amlodipine|swamlo|calchek|asomex|bisoprolol|bisotrol|atenolol|atenelo|captopril|methyldopa|nifedipine|losartan|presartan|presatan|candesartan|candez|aderan|atorvastatin|atorem|clopidogrel|clopact|frusemide|furosemide|\bascard\b|aspirin 75"),
    # --- antibiotics / anti-infectives -------------------------------------
    ("antibiotics", r"amoxycillin|amoxicillin|amoxycilin|ampiclox|ampicillin|clavam|clav\b|\d+clav|cledomox|koact|redmentin|indiclav|alphaclav|myclav|sportclav|spotclav|ciprofloxacin|\bcipro\b|ciprokant|cirpokant|zindolin|ofloxacin|\bpinox\b|\btoflox\b|norfloxacin|\bnor ?-?t\b|levofloxacin|\blevoz\b|azithromycin|\bazilin\b|\bazecure\b|\bazithraa\b|\bazuma\b|\bzaha\b|\bagycin\b|erythromycin|clarithromycin|\bclarie\b|\bclaranta\b|clindamycin|\btidact\b|doxycycline|\bdoxy\b|tetracycline|chloramphenicol|comycetin|\bprinaton\b|co-trimoxazole|cotrimoxazole|contrimoxazole|prinatrim|metronidazole|betrozole|prinalyn|tinidazole|ornidazole|\borgly\b|cefixime|cefibac|\bsanix\b|fixinet|inofix|\bc-tax\b|cefuroxime|auxtocef|cefadroxil|\bdrox\b|cephalexin|ceftriaxone|\baksone\b|flucloxacillin|flucamox|\bpen v\b|nitrofurantoin|\bnormax\b|nystatin|fluconazole|\bzocon\b|fluderm|flucoz|itraconazole|canditral|griseofulvin|terbinafine|binafin|ketoconazole|clotrimazole|miconazole|acyclovir|aciclovir|mupirocin|\bnetazox\b|\bfurazole\b|\bazitromycin\b"),
    # --- vitamins ----------------------------------------------------------
    ("vitamins-supplements", r"vitamin|\bbiocee\b|\bbecoshel\b|multivitamin|vitacap|\bvitamaks\b|\bdayvit\b|\bzenegy\b|\bnat b\b|\bnat d\b|\bginsomin\b|fish oil|\bmega\b|folic|\bferolic\b|\bferroton\b|\bglobin\b|\bhematon\b|\bhermovit\b|\bappet plus\b|\bappdly\b|\bgrowmin\b|\bgrwmin\b|\bzn vital\b|\bzinc\b|calcium|calcivita|caalvita|\bscotts\b|\bii-? ?care\b|livolin|\bfenza\b|neuro support|\bneuro ?-? ?forte\b|\bosteomin\b|joint support|21st century|\bglucose powder\b"),
    # --- mother & baby ------------------------------------------------------
    ("mother-baby", r"\bbaby\b|\bgripe water\b|zincast|\binfant\b|\bpaediatric\b|\bpead\b|\bjunior\b|\bnappy\b|\bdiaper\b"),
    # --- personal care -------------------------------------------------------
    ("personal-care", r"olive oil|castor oil|\bemami\b|\bopele\b|\bwater guard\b"),
]


def detect_category(raw: str) -> str | None:
    low = raw.lower()
    for slug, pattern in CATEGORY_RULES:
        if re.search(pattern, low):
            return slug
    return None


# --------------------------------------------------------------------------
# Manual overrides — row number (1-based, matching the spreadsheet data rows)
# to explicit values. Used where the rules cannot infer safely.
# --------------------------------------------------------------------------

OVERRIDES: dict[int, dict] = {}

try:
    _raw_overrides = json.loads((ROOT / "scripts" / "overrides.json").read_text())
    OVERRIDES = {
        int(k): v for k, v in _raw_overrides.items() if not k.startswith("_")
    }
except FileNotFoundError:
    pass


# --------------------------------------------------------------------------
# Brand detection
# --------------------------------------------------------------------------

BRANDS = {
    "gsk": "GSK", "cipla": "Cipla", "ajanta": "Ajanta", "shelys": "Shelys",
    "shels": "Shelys", "elys": "Elys", "ellys": "Elys", "zenufa": "Zenufa",
    "cosmol": "Cosmol", "cyprus": "Cyprus", "lincolin": "Lincolin",
    "remedica": "Remedica", "cadila": "Cadila", "ipca": "IPCA",
    "sedico": "Sedico", "eipico": "EIPICO", "21st century": "21st Century",
    "hovid": "Hovid", "sandox": "Sandoz", "denk": "Denk",
    "schering": "Schering", "bells": "Bells", "prince": "Prince",
    "s kant": "S. Kant", "s-kant": "S. Kant", "skant": "S. Kant",
    "l/a": "Laboratory & Allied", "akriti": "Akriti", "unique": "Unique",
    "heuer": "Heuer", "contempo": "Contempo", "durex": "Durex",
    "trust": "Trust", "mepha": "Mepha",
    "emami": "Emami", "wells": "Wells",
}


def detect_brand(raw: str) -> str | None:
    low = raw.lower()
    for key, name in BRANDS.items():
        if re.search(r"\b" + re.escape(key) + r"\b", low):
            return name
    return None


# --------------------------------------------------------------------------
# Build
# --------------------------------------------------------------------------


def to_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, str):
        return float(value.replace(",", "").strip())
    return float(value)


# Well-known over-the-counter lines shown on the homepage. Spreadsheet rows.
FEATURED_ROWS = {
    59,   # Panadol Extra Tablets 100s
    87,   # Panadol Advance Tablets
    376,  # ORS Sachet
    16,   # Vitamin C & Zinc Sugar Free Biocee 20s
    563,  # Gaviscon Double Action Liquid 150 ml
    222,  # Sensodyne Original 50 ml
    423,  # Strepsils Tablets 24s
    677,  # Zecuf Syrup 100 ml
    439,  # Deep Heat Cream 35 g
    570,  # Scotts Original 100 ml
    218,  # Panadol Baby Infant Suspension 100 ml
    193,  # Cotton Wool 50 g
}


def low_stock_threshold(stock: int) -> int:
    if stock <= 10:
        return 3
    if stock <= 50:
        return 10
    if stock <= 200:
        return 20
    return 50


def main() -> int:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.worksheets[0]

    raw_rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name is None or not str(name).strip():
            continue
        raw_rows.append({
            "row": len(raw_rows) + 1,
            "sheet_row": r,
            "internal_name": str(name).strip(),
            "cost": to_number(ws.cell(r, 2).value),
            "price": to_number(ws.cell(r, 3).value),
            "stock": int(to_number(ws.cell(r, 7).value)),
        })

    # merge byte-identical duplicate lines (same name, price and stock)
    merged: dict[tuple, dict] = {}
    order: list[tuple] = []
    duplicates_merged = []
    for row in raw_rows:
        key = (row["internal_name"], row["price"], row["stock"])
        if key in merged:
            merged[key]["stock"] += row["stock"]
            merged[key]["merged_rows"].append(row["row"])
            duplicates_merged.append(row)
            continue
        row["merged_rows"] = [row["row"]]
        merged[key] = row
        order.append(key)
    rows = [merged[k] for k in order]

    products = []
    slugs: dict[str, int] = {}
    sku_counters: Counter = Counter()
    uncategorised = []
    normalised_names = defaultdict(list)
    renamed = []
    review_notes = []

    for row in rows:
        raw = row["internal_name"]
        override = OVERRIDES.get(row["row"], {})

        derived = clean_name(raw)
        name = override.get("name") or derived
        if override.get("name") and override["name"] != derived:
            renamed.append((row, derived, override["name"]))
        if override.get("review"):
            review_notes.append((row, override["review"]))
        # Match rules against both the raw spreadsheet text and the cleaned
        # name, so corrected spellings ("COTTOL WOOL" -> "Cotton Wool") match.
        searchable = f"{raw} || {name}"

        category = override.get("category") or detect_category(searchable) or "general-medicines"
        generic = override.get("generic", detect_generic(searchable))
        form = override.get("form") or detect_form(searchable)
        pack = override.get("pack", detect_pack(name, form))
        brand = override.get("brand", detect_brand(searchable))

        if not override.get("category") and not detect_category(searchable):
            uncategorised.append(row)

        base = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
        slug = base
        if slug in slugs:
            slugs[base] += 1
            slug = f"{base}-{slugs[base]}"
        else:
            slugs[base] = 1

        code = CATEGORY_CODES[category]
        sku_counters[code] += 1
        sku = f"MED-{code}-{sku_counters[code]:03d}"

        keywords = set()
        if generic:
            keywords.add(generic.lower())
        if form:
            keywords.add(form.lower())
        for token in re.split(r"[^A-Za-z]+", raw):
            if len(token) >= 4:
                keywords.add(token.lower())
        keyword_list = sorted(keywords)[:12]

        short = None
        if generic and form:
            short = f"{generic} {form.lower()}."
        elif generic:
            short = f"{generic}."

        products.append({
            "row": row["row"],
            "merged_rows": row["merged_rows"],
            "name": name,
            "slug": slug,
            "internal_name": raw,
            "generic_name": generic,
            "brand": brand,
            "category": category,
            "sku": sku,
            "pack_size": pack,
            "short_description": short,
            "keywords": keyword_list,
            "price": row["price"],
            # NOTE: the spreadsheet's Cost column is deliberately NOT carried
            # through. It is the pharmacy's buying price — commercially
            # sensitive, and this repository is public.
            "stock_quantity": row["stock"],
            "low_stock_threshold": low_stock_threshold(row["stock"]),
            "is_featured": row["row"] in FEATURED_ROWS,
        })

        normalised_names[re.sub(r"[^a-z0-9]", "", name.lower())].append(products[-1])

    OUT_JSON.write_text(json.dumps({
        "categories": [
            {"slug": s, "name": n, "code": c, "description": d}
            for s, n, c, d in CATEGORIES
        ],
        "products": products,
    }, indent=1, ensure_ascii=False))

    # ---- report -----------------------------------------------------------
    by_cat = Counter(p["category"] for p in products)
    near_dupes = {k: v for k, v in normalised_names.items() if len(v) > 1}

    lines = ["# Catalogue build report", ""]
    lines.append(f"- Spreadsheet rows read: **{len(raw_rows)}**")
    lines.append(f"- Identical duplicate lines merged: **{len(duplicates_merged)}**")
    lines.append(f"- Products written: **{len(products)}**")
    lines.append(f"- Total stock units: **{sum(p['stock_quantity'] for p in products):,}**")
    lines.append("")
    lines.append("## Products per category")
    lines.append("")
    lines.append("| Category | Products |")
    lines.append("| --- | ---: |")
    for slug, name, _c, _d in CATEGORIES:
        lines.append(f"| {name} | {by_cat.get(slug, 0)} |")
    lines.append("")

    if duplicates_merged:
        lines.append("## Merged duplicate lines")
        lines.append("")
        lines.append("Identical name, price and stock on two rows — stock was added together.")
        lines.append("")
        for d in duplicates_merged:
            lines.append(f"- Row {d['row']}: `{d['internal_name']}` (+{d['stock']} units)")
        lines.append("")

    if near_dupes:
        lines.append("## Possible duplicate products — please review")
        lines.append("")
        lines.append("Same product name after clean-up but different price or stock. "
                     "Both were kept; edit or archive one in the admin dashboard if needed.")
        lines.append("")
        for _key, group in sorted(near_dupes.items()):
            lines.append(f"- **{group[0]['name']}**")
            for p in group:
                lines.append(
                    f"  - row {p['row']}: `{p['internal_name']}` — "
                    f"TZS {p['price']:,.0f}, stock {p['stock_quantity']}"
                )
        lines.append("")

    if renamed:
        lines.append("## Manual name corrections — please confirm")
        lines.append("")
        lines.append("Rows where the spreadsheet unit looked wrong by a factor of a "
                     "thousand (a tablet strength written in `gm`, or a cream tube "
                     "size written in `mg`) or the text was garbled. Prices and stock "
                     "were **not** touched.")
        lines.append("")
        lines.append("| Row | Spreadsheet name | Published as |")
        lines.append("| ---: | --- | --- |")
        for row, derived, final in renamed:
            lines.append(f"| {row['row']} | `{row['internal_name']}` | {final} |")
        lines.append("")

    if review_notes:
        lines.append("## Needs a pharmacist's eye")
        lines.append("")
        for row, note in review_notes:
            lines.append(f"- **row {row['row']}** `{row['internal_name']}` — {note}")
        lines.append("")

    if uncategorised:
        lines.append("## Fell through to General Medicines")
        lines.append("")
        lines.append("No category rule matched, so these were filed under "
                     "General Medicines. Re-categorise any of them in "
                     "Admin → Products → Edit.")
        lines.append("")
        for row in uncategorised:
            lines.append(f"- row {row['row']}: `{row['internal_name']}`")
        lines.append("")

    OUT_REVIEW.write_text("\n".join(lines))

    print(f"rows={len(raw_rows)} merged={len(duplicates_merged)} products={len(products)}")
    print("uncategorised:", len(uncategorised))
    for slug, name, _c, _d in CATEGORIES:
        print(f"  {by_cat.get(slug,0):4d}  {name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
